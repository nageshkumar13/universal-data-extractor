import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

try:
    import pandas as pd
except ImportError:
    pd = None


class Exporter:
    def to_csv(self, records: list[dict], output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if pd is not None:
            dataframe = pd.DataFrame(records)
            dataframe.to_csv(output_path, index=False)
            return output_path

        fieldnames = list(records[0].keys()) if records else []
        with output_path.open("w", encoding="utf-8", newline="") as file_handle:
            writer = csv.DictWriter(file_handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(records)

        return output_path

    def to_json(self, records: list[dict], output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if pd is not None:
            dataframe = pd.DataFrame(records)
            dataframe.to_json(
                output_path,
                orient="records",
                indent=4,
                force_ascii=False,
            )
            return output_path

        with output_path.open("w", encoding="utf-8") as file_handle:
            json.dump(records, file_handle, indent=4, ensure_ascii=False)

        return output_path

    def to_excel(self, records: list[dict], output_path: Path) -> Path:
        """Export records to an Excel .xlsx file."""
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.freeze_panes = "A2"

        headers = list(records[0].keys()) if records else []
        if headers:
            worksheet.append(headers)
            for cell in worksheet[1]:
                cell.font = Font(bold=True)

            for record in records:
                worksheet.append([record.get(header, "") for header in headers])

            for index, header in enumerate(headers, start=1):
                values = [header]
                values.extend("" if record.get(header) is None else str(record.get(header)) for record in records)
                width = min(max(len(value) for value in values) + 2, 50)
                worksheet.column_dimensions[get_column_letter(index)].width = width

        workbook.save(output_path)
        return output_path
